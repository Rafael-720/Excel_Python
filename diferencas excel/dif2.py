
import pandas as pd
from openpyxl import Workbook

# Função para comparar duas linhas, considerando campos vazios
def compare_rows(row1, row2):
    changed_fields = []
    for col in row1.index:
        val1 = row1[col]
        val2 = row2[col]
        if pd.isna(val1) and pd.isna(val2):
            continue
        if val1 != val2:
            changed_fields.append(col)
    return changed_fields

# Carregar as planilhas para DataFrames
df1 = pd.read_excel('C:/Users/Rafael/Desktop/tempM/planilha1.xlsx')
df2 = pd.read_excel('C:/Users/Rafael/Desktop/tempM/planilha2.xlsx')

# Ordenar as planilhas pela coluna "descricao"
df1.sort_values(by='descricao', inplace=True)
df2.sort_values(by='descricao', inplace=True)

# Encontrar descrições adicionadas e removidas
descricao_added = df2[~df2['descricao'].isin(df1['descricao'])]['descricao'].tolist()
descricao_removed = df1[~df1['descricao'].isin(df2['descricao'])]['descricao'].tolist()

# Encontrar descrições que estão em ambas as planilhas para verificar mudanças
df_common1 = df1[df1['descricao'].isin(df2['descricao'])].sort_values(by='descricao')
df_common2 = df2[df2['descricao'].isin(df1['descricao'])].sort_values(by='descricao')

# Comparar as linhas com descrições comuns e identificar campos alterados (corrigindo para campos vazios)
corrected_changes = []
for _, row1 in df_common1.iterrows():
    row2 = df_common2[df_common2['descricao'] == row1['descricao']].iloc[0]
    changed_fields = compare_rows(row1, row2)
    if changed_fields:
        # Remover 'descricao' da lista de campos alterados, já que estamos usando isso como chave
        if 'descricao' in changed_fields:
            changed_fields.remove('descricao')
        corrected_changes.append({"descricao": row1['descricao'], "changed_fields": changed_fields})

# Criar um novo Workbook do Excel para armazenar as correções
wb_corrected = Workbook()

# Adicionar uma planilha para descrições adicionadas
ws_added = wb_corrected.create_sheet("Adicionadas")
ws_added.append(["Descrição"])
for desc in descricao_added:
    ws_added.append([desc])

# Adicionar uma planilha para descrições removidas
ws_removed = wb_corrected.create_sheet("Removidas")
ws_removed.append(["Descrição"])
for desc in descricao_removed:
    ws_removed.append([desc])

# Adicionar uma planilha para descrições alteradas (corrigidas)
ws_changed_corrected = wb_corrected.create_sheet("Alteradas_Corrigidas")
ws_changed_corrected.append(["Descrição", "Campos Alterados"])
for change in corrected_changes:
    ws_changed_corrected.append([change['descricao'], ", ".join(change['changed_fields'])])

# Remover a planilha padrão em branco criada
if 'Sheet' in wb_corrected.sheetnames:
    wb_corrected.remove(wb_corrected['Sheet'])

# Salvar o Workbook corrigido
wb_corrected.save('C:/Users/Rafael/Desktop/tempM/arquivo_corrigido.xlsx')
